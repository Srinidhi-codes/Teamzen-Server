"""Smart Excel/CSV import: parse → map columns (heuristic + AI) → preview → commit."""

from __future__ import annotations

import csv
import io
import json
import logging
import re
import secrets
import string
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

logger = logging.getLogger(__name__)

# Target fields HR can map spreadsheet columns onto
IMPORT_TARGET_FIELDS: list[dict[str, str]] = [
    {"key": "email", "label": "Email", "required": "true"},
    {"key": "first_name", "label": "First name", "required": "false"},
    {"key": "last_name", "label": "Last name", "required": "false"},
    {"key": "employee_id", "label": "Employee ID", "required": "false"},
    {"key": "phone_number", "label": "Phone", "required": "false"},
    {"key": "pan_number", "label": "PAN", "required": "false"},
    {"key": "aadhar_number", "label": "Aadhaar", "required": "false"},
    {"key": "uan_number", "label": "UAN", "required": "false"},
    {"key": "bank_account_number", "label": "Bank account", "required": "false"},
    {"key": "bank_ifsc_code", "label": "IFSC", "required": "false"},
    {"key": "date_of_joining", "label": "Date of joining", "required": "false"},
    {"key": "date_of_birth", "label": "Date of birth", "required": "false"},
    {"key": "gender", "label": "Gender", "required": "false"},
    {"key": "department", "label": "Department (name)", "required": "false"},
    {"key": "designation", "label": "Designation (name)", "required": "false"},
    {"key": "annual_ctc", "label": "Annual CTC", "required": "false"},
    {"key": "employment_type", "label": "Employment type", "required": "false"},
]

TARGET_KEYS = {f["key"] for f in IMPORT_TARGET_FIELDS}

# header fragment → target field (lowercased, stripped)
_HEADER_ALIASES: dict[str, str] = {
    "email": "email",
    "e-mail": "email",
    "work email": "email",
    "official email": "email",
    "mail id": "email",
    "email id": "email",
    "email address": "email",
    "first name": "first_name",
    "firstname": "first_name",
    "given name": "first_name",
    "last name": "last_name",
    "lastname": "last_name",
    "surname": "last_name",
    "employee name": "first_name",  # split later if needed
    "name": "first_name",
    "full name": "first_name",
    "emp id": "employee_id",
    "emp code": "employee_id",
    "employee id": "employee_id",
    "employee code": "employee_id",
    "employee no": "employee_id",
    "employee number": "employee_id",
    "staff id": "employee_id",
    "phone": "phone_number",
    "mobile": "phone_number",
    "phone number": "phone_number",
    "mobile number": "phone_number",
    "contact": "phone_number",
    "pan": "pan_number",
    "pan number": "pan_number",
    "pan no": "pan_number",
    "aadhaar": "aadhar_number",
    "aadhar": "aadhar_number",
    "aadhaar number": "aadhar_number",
    "aadhar number": "aadhar_number",
    "uan": "uan_number",
    "uan number": "uan_number",
    "bank account": "bank_account_number",
    "account number": "bank_account_number",
    "account no": "bank_account_number",
    "bank a/c": "bank_account_number",
    "bank account number": "bank_account_number",
    "ifsc": "bank_ifsc_code",
    "ifsc code": "bank_ifsc_code",
    "date of joining": "date_of_joining",
    "doj": "date_of_joining",
    "joining date": "date_of_joining",
    "join date": "date_of_joining",
    "date of birth": "date_of_birth",
    "dob": "date_of_birth",
    "birth date": "date_of_birth",
    "gender": "gender",
    "sex": "gender",
    "department": "department",
    "dept": "department",
    "designation": "designation",
    "title": "designation",
    "job title": "designation",
    "role": "designation",
    "ctc": "annual_ctc",
    "annual ctc": "annual_ctc",
    "yearly ctc": "annual_ctc",
    "cost to company": "annual_ctc",
    "salary": "annual_ctc",
    "annual salary": "annual_ctc",
    "employment type": "employment_type",
    "emp type": "employment_type",
    "type": "employment_type",
}


def _norm_header(h: str) -> str:
    return re.sub(r"\s+", " ", (h or "").strip().lower())


def parse_tabular_file(file_bytes: bytes, file_name: str) -> tuple[str, list[str], list[dict]]:
    """
    Parse CSV or Excel into (source_type, headers, rows-as-dicts).
    Rows use original header keys.
    """
    name = (file_name or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        return "xlsx", *_parse_xlsx(file_bytes)
    if name.endswith(".xls"):
        # openpyxl does not read legacy .xls — try as xlsx anyway then CSV
        try:
            return "xlsx", *_parse_xlsx(file_bytes)
        except Exception:
            return "csv", *_parse_csv(file_bytes)
    return "csv", *_parse_csv(file_bytes)


def _parse_csv(file_bytes: bytes) -> tuple[list[str], list[dict]]:
    text = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = file_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = file_bytes.decode("utf-8", errors="replace")

    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("CSV has no header row")
    headers = [str(h).strip() for h in reader.fieldnames if h is not None and str(h).strip()]
    rows: list[dict] = []
    for raw in reader:
        row = {h: (raw.get(h) or "").strip() if isinstance(raw.get(h), str) else raw.get(h)
               for h in headers}
        if any(str(v or "").strip() for v in row.values()):
            rows.append(row)
    return headers, rows


def _parse_xlsx(file_bytes: bytes) -> tuple[list[str], list[dict]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ValueError(
            "Excel support requires openpyxl. Install it or upload a CSV instead."
        ) from e

    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ValueError("Excel sheet is empty")

    headers: list[str] = []
    for i, cell in enumerate(header_row):
        if cell is None or str(cell).strip() == "":
            headers.append(f"Column_{i + 1}")
        else:
            headers.append(str(cell).strip())

    # Deduplicate headers
    seen: dict[str, int] = {}
    deduped: list[str] = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            deduped.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            deduped.append(h)
    headers = deduped

    rows: list[dict] = []
    for values in rows_iter:
        row: dict[str, Any] = {}
        empty = True
        for i, h in enumerate(headers):
            val = values[i] if i < len(values) else None
            if val is None:
                row[h] = ""
            elif isinstance(val, datetime):
                row[h] = val.date().isoformat()
                empty = False
            elif isinstance(val, date):
                row[h] = val.isoformat()
                empty = False
            else:
                s = str(val).strip()
                row[h] = s
                if s:
                    empty = False
        if not empty:
            rows.append(row)
    wb.close()
    return headers, rows


def heuristic_column_mapping(headers: list[str]) -> tuple[dict[str, str], dict[str, float]]:
    """Map spreadsheet headers → target fields using aliases. Confidence 0–1."""
    mapping: dict[str, str] = {}
    confidence: dict[str, float] = {}
    used_targets: set[str] = set()

    for header in headers:
        key = _norm_header(header)
        target = _HEADER_ALIASES.get(key)
        conf = 0.95 if target else 0.0
        if not target:
            # fuzzy: contains
            for alias, t in _HEADER_ALIASES.items():
                if alias in key or key in alias:
                    target = t
                    conf = 0.7
                    break
        if target and target not in used_targets:
            mapping[header] = target
            confidence[header] = conf
            used_targets.add(target)
        else:
            mapping[header] = ""
            confidence[header] = 0.0
    return mapping, confidence


def ai_refine_column_mapping(
    organization_id: int,
    headers: list[str],
    sample_rows: list[dict],
    current_mapping: dict[str, str] | None = None,
) -> tuple[dict[str, str], dict[str, float]]:
    """Ask org LLM to improve column mapping. Falls back to heuristic on failure."""
    base, base_conf = heuristic_column_mapping(headers)
    if current_mapping:
        for h, t in current_mapping.items():
            if h in base and t in TARGET_KEYS:
                base[h] = t
                base_conf[h] = max(base_conf.get(h, 0), 0.8)

    try:
        from ai_engine.graph import get_llm
        from ai_engine.views import _normalize_llm_content
        from langchain_core.messages import HumanMessage, SystemMessage

        targets = ", ".join(sorted(TARGET_KEYS))
        sample = json.dumps(sample_rows[:5], default=str)[:4000]
        system = (
            "You map spreadsheet columns to Teamzen employee/payroll fields. "
            f"Valid target keys: {targets}. "
            "Return ONLY a JSON object: "
            '{"mapping": {"Header Name": "target_key_or_empty"}, '
            '"confidence": {"Header Name": 0.0_to_1.0}}. '
            "Use empty string to skip a column. email is required if present. "
            "No markdown."
        )
        user_msg = (
            f"Headers: {json.dumps(headers)}\n"
            f"Current guess: {json.dumps(base)}\n"
            f"Sample rows: {sample}"
        )
        llm = get_llm(organization_id)
        result = llm.invoke(
            [SystemMessage(content=system), HumanMessage(content=user_msg)]
        )
        text = _normalize_llm_content(getattr(result, "content", result)).strip()
        data = _extract_json_object(text)
        raw_map = data.get("mapping") if isinstance(data, dict) else None
        raw_conf = data.get("confidence") if isinstance(data, dict) else None
        if not isinstance(raw_map, dict):
            return base, base_conf

        used: set[str] = set()
        out: dict[str, str] = {}
        conf_out: dict[str, float] = {}
        for h in headers:
            t = str(raw_map.get(h, base.get(h, "")) or "").strip()
            if t not in TARGET_KEYS:
                t = ""
            if t and t in used:
                t = ""
            if t:
                used.add(t)
            out[h] = t
            try:
                c = float((raw_conf or {}).get(h, base_conf.get(h, 0.5)))
            except (TypeError, ValueError):
                c = base_conf.get(h, 0.5)
            conf_out[h] = max(0.0, min(1.0, c))
        return out, conf_out
    except Exception:
        logger.exception("AI column mapping failed; using heuristic")
        return base, base_conf


def _extract_json_object(text: str) -> dict:
    raw = (text or "").strip()
    if raw.startswith("```"):
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
    try:
        data = json.loads(raw)
        if isinstance(data, dict):
            return data
    except json.JSONDecodeError:
        pass
    match = re.search(r"\{[\s\S]*\}", raw)
    if match:
        try:
            data = json.loads(match.group(0))
            if isinstance(data, dict):
                return data
        except json.JSONDecodeError:
            pass
    return {}


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%d.%m.%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s.replace("Z", "")).date()
    except ValueError:
        return None


def _parse_ctc(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    s = str(value).strip().replace(",", "").replace("₹", "").replace("Rs", "").replace("INR", "")
    s = re.sub(r"[^\d.]", "", s)
    if not s:
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _split_name(full: str) -> tuple[str, str]:
    parts = full.strip().split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def row_to_record(row: dict, mapping: dict[str, str]) -> dict[str, Any]:
    """Apply column mapping to one spreadsheet row → normalized record."""
    record: dict[str, Any] = {}
    for header, target in mapping.items():
        if not target or target not in TARGET_KEYS:
            continue
        raw = row.get(header, "")
        if raw is None:
            continue
        val = str(raw).strip() if not isinstance(raw, (int, float, Decimal, date, datetime)) else raw
        if val == "" or val is None:
            continue
        if target in ("date_of_joining", "date_of_birth"):
            d = _parse_date(val)
            if d:
                record[target] = d.isoformat()
        elif target == "annual_ctc":
            ctc = _parse_ctc(val)
            if ctc is not None:
                record[target] = str(ctc)
        elif target == "employment_type":
            et = str(val).lower().replace(" ", "_").replace("-", "_")
            if "intern" in et:
                record[target] = "intern"
            elif "contract" in et or "consultant" in et:
                record[target] = "contract"
            else:
                record[target] = "full_time"
        elif target == "gender":
            g = str(val).lower()
            if g.startswith("m"):
                record[target] = "male"
            elif g.startswith("f"):
                record[target] = "female"
            else:
                record[target] = g[:20]
        else:
            record[target] = str(val).strip()

    # If only first_name looks like a full name and last_name missing
    if record.get("first_name") and not record.get("last_name"):
        fn, ln = _split_name(record["first_name"])
        if ln:
            record["first_name"] = fn
            record["last_name"] = ln
    return record


def build_preview(organization, rows: list[dict], mapping: dict[str, str]) -> dict:
    """Validate mapped rows against org users; return create/update/error buckets."""
    from django.contrib.auth import get_user_model

    User = get_user_model()
    existing_emails = {
        e.lower(): uid
        for uid, e in User.objects.filter(organization=organization)
        .exclude(email__isnull=True)
        .exclude(email="")
        .values_list("id", "email")
    }
    existing_emp_ids = set(
        User.objects.filter(organization=organization)
        .exclude(employee_id__isnull=True)
        .exclude(employee_id="")
        .values_list("employee_id", flat=True)
    )

    to_create: list[dict] = []
    to_update: list[dict] = []
    errors: list[dict] = []
    seen_emails: set[str] = set()

    has_email_col = "email" in mapping.values()

    for idx, row in enumerate(rows, start=2):  # 1-based + header
        rec = row_to_record(row, mapping)
        email = (rec.get("email") or "").lower().strip()
        issues: list[str] = []

        if not has_email_col:
            issues.append("No email column mapped")
        elif not email:
            issues.append("Missing email")
        elif "@" not in email:
            issues.append("Invalid email")
        elif email in seen_emails:
            issues.append("Duplicate email in file")

        if email:
            seen_emails.add(email)

        emp_id = rec.get("employee_id")
        if emp_id and emp_id in existing_emp_ids and email not in existing_emails:
            # emp_id taken by someone else — warn but allow create without emp_id override later
            issues.append(f"Employee ID {emp_id} already exists in org")

        entry = {
            "row": idx,
            "email": email or None,
            "first_name": rec.get("first_name") or "",
            "last_name": rec.get("last_name") or "",
            "employee_id": emp_id or None,
            "annual_ctc": rec.get("annual_ctc"),
            "department": rec.get("department"),
            "designation": rec.get("designation"),
            "record": rec,
            "issues": issues,
        }

        if issues and (not email or "Missing email" in issues or "Invalid email" in issues
                       or "No email column mapped" in issues or "Duplicate email in file" in issues):
            errors.append(entry)
            continue

        if email in existing_emails:
            entry["user_id"] = existing_emails[email]
            entry["action"] = "update"
            # non-blocking issues stay as warnings
            to_update.append(entry)
        else:
            entry["action"] = "create"
            to_create.append(entry)

    return {
        "total_rows": len(rows),
        "create_count": len(to_create),
        "update_count": len(to_update),
        "error_count": len(errors),
        "to_create": to_create[:100],
        "to_update": to_update[:100],
        "errors": errors[:100],
        "truncated": len(to_create) > 100 or len(to_update) > 100 or len(errors) > 100,
    }


def _temp_password(length: int = 12) -> str:
    alphabet = string.ascii_letters + string.digits
    return "".join(secrets.choice(alphabet) for _ in range(length))


def _get_or_create_dept(organization, name: str):
    from organizations.models import Department

    name = (name or "").strip()
    if not name:
        return None
    dept = Department.objects.filter(
        organization=organization, name__iexact=name
    ).first()
    if dept:
        return dept
    return Department.objects.create(organization=organization, name=name[:100])


def _get_or_create_designation(organization, name: str):
    from organizations.models import Designation

    name = (name or "").strip()
    if not name:
        return None
    des = Designation.objects.filter(
        organization=organization, name__iexact=name
    ).first()
    if des:
        return des
    return Designation.objects.create(organization=organization, name=name[:100])


def _ensure_import_salary_structure(organization):
    """Return a basic salary structure for imported CTC assignments."""
    from payroll.setup_services import ensure_default_salary_structure

    return ensure_default_salary_structure(organization)


def _assign_ctc(user, organization, annual_ctc: str | Decimal, effective_from: date | None = None):
    from payroll.models import EmployeeSalaryStructure

    ctc = _parse_ctc(annual_ctc)
    if ctc is None or ctc <= 0:
        return None
    structure = _ensure_import_salary_structure(organization)
    EmployeeSalaryStructure.objects.filter(user=user, is_active=True).update(is_active=False)
    return EmployeeSalaryStructure.objects.create(
        user=user,
        salary_structure=structure,
        annual_ctc=ctc,
        effective_from=effective_from or date.today(),
        is_active=True,
    )


def commit_import(
    organization,
    rows: list[dict],
    mapping: dict[str, str],
    *,
    update_existing: bool = True,
    assign_ctc: bool = True,
    send_welcome: bool = False,
    actor=None,
) -> dict:
    """Create/update users (and optional CTC) from mapped rows."""
    from django.contrib.auth import get_user_model
    from django.db import transaction

    User = get_user_model()
    preview = build_preview(organization, rows, mapping)

    created = 0
    updated = 0
    ctc_assigned = 0
    failed: list[dict] = []

    existing_by_email = {
        u.email.lower(): u
        for u in User.objects.filter(organization=organization)
        if u.email
    }

    with transaction.atomic():
        for idx, row in enumerate(rows, start=2):
            rec = row_to_record(row, mapping)
            email = (rec.get("email") or "").lower().strip()
            if not email or "@" not in email:
                failed.append({"row": idx, "error": "Missing or invalid email"})
                continue

            try:
                user = existing_by_email.get(email)
                is_new = user is None

                if is_new:
                    password = _temp_password()
                    user = User.objects.create_user(
                        email=email,
                        username=email,
                        password=password,
                        first_name=rec.get("first_name") or email.split("@")[0],
                        last_name=rec.get("last_name") or "",
                        role="employee",
                        employment_type=rec.get("employment_type") or "full_time",
                        is_active=True,
                        organization=organization,
                    )
                    existing_by_email[email] = user
                    created += 1

                    if send_welcome and actor:
                        try:
                            from notifications.utils import notify_user

                            notify_user(
                                recipient_id=user.id,
                                verb="Welcome to Teamzen",
                                message=(
                                    f"Your Teamzen account has been created. "
                                    f"Sign in with {user.email}."
                                ),
                                actor_id=actor.id,
                                target_type="Welcome",
                                target_id=str(user.id),
                                level="personal",
                                notification_type="BOTH",
                                extra_context={"temp_password": password},
                            )
                        except Exception:
                            logger.exception("Welcome email failed for %s", email)
                else:
                    if not update_existing:
                        continue
                    updated += 1

                # Shared field updates
                for field in (
                    "first_name",
                    "last_name",
                    "phone_number",
                    "pan_number",
                    "aadhar_number",
                    "uan_number",
                    "bank_account_number",
                    "bank_ifsc_code",
                    "gender",
                    "employment_type",
                ):
                    if rec.get(field):
                        setattr(user, field, rec[field])

                if rec.get("employee_id"):
                    clash = (
                        User.objects.filter(employee_id=rec["employee_id"])
                        .exclude(id=user.id)
                        .exists()
                    )
                    if not clash:
                        user.employee_id = rec["employee_id"]

                if rec.get("date_of_joining"):
                    d = _parse_date(rec["date_of_joining"])
                    if d:
                        user.date_of_joining = d
                if rec.get("date_of_birth"):
                    d = _parse_date(rec["date_of_birth"])
                    if d:
                        user.date_of_birth = d

                if rec.get("department"):
                    dept = _get_or_create_dept(organization, rec["department"])
                    if dept:
                        user.department = dept
                if rec.get("designation"):
                    des = _get_or_create_designation(organization, rec["designation"])
                    if des:
                        user.designation = des

                user.save()

                if assign_ctc and rec.get("annual_ctc"):
                    eff = _parse_date(rec.get("date_of_joining")) or date.today()
                    if _assign_ctc(user, organization, rec["annual_ctc"], effective_from=eff):
                        ctc_assigned += 1

            except Exception as e:
                logger.exception("Import row %s failed", idx)
                failed.append({"row": idx, "email": email, "error": str(e)})

    return {
        "created": created,
        "updated": updated,
        "ctc_assigned": ctc_assigned,
        "failed_count": len(failed),
        "failed": failed[:50],
        "preview_summary": {
            "create_count": preview["create_count"],
            "update_count": preview["update_count"],
            "error_count": preview["error_count"],
        },
    }
